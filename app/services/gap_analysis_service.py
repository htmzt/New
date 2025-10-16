from typing import Dict, Any, List, Optional
from sqlalchemy.orm import Session
from sqlalchemy import text
from app.services.base_service import BaseService
from app.query import MERGED_DATA_QUERY  # Import the merge query
import pandas as pd
from io import BytesIO
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

class GapAnalysisService(BaseService):
    def __init__(self, db: Session):
        super().__init__(db)
    

    def get_gap_financial_summary_by_project(self, user_id: str, project_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """Get gap financial analysis summary in tabular format"""
        try:
            # Build base filter
            base_filter = "po.user_id = :user_id"
            params = {"user_id": user_id}
            
            # Add project filter if provided
            if project_name:
                base_filter += " AND po.project_name ILIKE :project_name"
                params["project_name"] = f"%{project_name}%"
            
            # Direct SQL query for financial summary
            financial_summary_query = f"""
            WITH base_data AS (
                {MERGED_DATA_QUERY.format(base_filter=base_filter)}
            ),
            filtered_data AS (
                SELECT *,
                    CASE
                        WHEN status IN ('Pending AC80%', 'Pending ACPAC') AND ac_date IS NULL THEN 'AC_PENDING'
                        WHEN status = 'Pending PAC20%' AND pac_date IS NULL AND ac_date IS NOT NULL THEN 'PAC_PENDING'
                        WHEN status = 'CLOSED' THEN 'COMPLETED'
                        WHEN status = 'CANCELLED' THEN 'CANCELLED'
                        ELSE 'OTHER'
                    END AS gap_type
                FROM base_data
                WHERE publish_date IS NOT NULL
                AND status != 'CANCELLED'
            ),
            project_summary AS (
                SELECT 
                    project_name,
                    SUM(COALESCE(line_amount, 0)) as total_po_received,
                    SUM(CASE WHEN gap_type = 'AC_PENDING' THEN COALESCE(line_amount, 0) ELSE 0 END) as gap_po_nok_ac_nok,
                    SUM(CASE WHEN gap_type = 'PAC_PENDING' THEN COALESCE(line_amount, 0) ELSE 0 END) as gap_ac_ok_pac_nok,
                    SUM(CASE WHEN gap_type IN ('AC_PENDING', 'PAC_PENDING') THEN COALESCE(line_amount, 0) ELSE 0 END) as total_gap_ac_pac
                FROM filtered_data
                WHERE project_name IS NOT NULL AND TRIM(project_name) != ''
                GROUP BY project_name
            )
            SELECT 
                project_name,
                ROUND(total_po_received, 2) as total_po_received,
                ROUND(gap_po_nok_ac_nok, 2) as gap_po_nok_ac_nok, 
                ROUND(gap_ac_ok_pac_nok, 2) as gap_ac_ok_pac_nok,
                ROUND(total_gap_ac_pac, 2) as total_gap_ac_pac,
                CASE 
                    WHEN total_po_received > 0 THEN 
                        ROUND((total_gap_ac_pac / total_po_received * 100), 0)
                    ELSE 0
                END as gap_percentage
            FROM project_summary
            ORDER BY total_po_received DESC
            """
            
            result = self.db.execute(text(financial_summary_query), params)
            data = result.fetchall()
            column_names = list(result.keys())
            
            # Convert to list of dictionaries with comma as decimal separator
            financial_summary_table = []
            total_po = 0
            total_ac_nok = 0
            total_pac_nok = 0
            total_gap = 0
            
            for row in data:
                row_dict = dict(zip(column_names, row))
                total_po += row_dict['total_po_received']
                total_ac_nok += row_dict['gap_po_nok_ac_nok']
                total_pac_nok += row_dict['gap_ac_ok_pac_nok']
                total_gap += row_dict['total_gap_ac_pac']

                gap_percentage = int(row_dict['gap_percentage'])

                financial_summary_table.append({
                    "project_name": row_dict['project_name'],
                    "total_po_received": str(row_dict['total_po_received']).replace('.', ','),
                    "gap_po_nok_ac_nok": str(row_dict['gap_po_nok_ac_nok']).replace('.', ','),
                    "gap_ac_ok_pac_nok": str(row_dict['gap_ac_ok_pac_nok']).replace('.', ','),
                    "total_gap_ac_pac": str(row_dict['total_gap_ac_pac']).replace('.', ','),
                    "gap_percentage": f"{int(row_dict['gap_percentage'])}%",
                    "completion_percentage": f"{100 - gap_percentage}%"
                })
            
            # Add total row
            total_gap_percentage = int(round((total_gap / total_po * 100), 0)) if total_po > 0 else 0
            financial_summary_table.append({
                "project_name": "TOTAL",
                "total_po_received": str(round(total_po, 2)).replace('.', ','),
                "gap_po_nok_ac_nok": str(round(total_ac_nok, 2)).replace('.', ','),
                "gap_ac_ok_pac_nok": str(round(total_pac_nok, 2)).replace('.', ','),
                "total_gap_ac_pac": str(round(total_gap, 2)).replace('.', ','),
                "gap_percentage": f"{total_gap_percentage}%",
                "completion_percentage": f"{100 - total_gap_percentage}%" 
            })
            
            return financial_summary_table
            
        except Exception as e:
            logger.error(f"Error getting gap financial summary: {str(e)}")
            raise

    def export_gap_financial_summary_to_excel(self, user_id: str, project_name: Optional[str] = None) -> bytes:
        """Export gap financial summary directly to Excel"""
        try:
            # Build base filter
            base_filter = "po.user_id = :user_id"
            params = {"user_id": user_id}
            
            # Add project filter if provided
            if project_name:
                base_filter += " AND po.project_name ILIKE :project_name"
                params["project_name"] = f"%{project_name}%"
            
            # Direct SQL query for Excel export (including totals)
            financial_summary_query = f"""
            WITH base_data AS (
                {MERGED_DATA_QUERY.format(base_filter=base_filter)}
            ),
            filtered_data AS (
                SELECT *,
                    CASE
                        WHEN status IN ('Pending AC80%', 'Pending ACPAC') AND ac_date IS NULL THEN 'AC_PENDING'
                        WHEN status = 'Pending PAC20%' AND pac_date IS NULL AND ac_date IS NOT NULL THEN 'PAC_PENDING'
                        WHEN status = 'CLOSED' THEN 'COMPLETED'
                        WHEN status = 'CANCELLED' THEN 'CANCELLED'
                        ELSE 'OTHER'
                    END AS gap_type
                FROM base_data
                WHERE publish_date IS NOT NULL
                AND status != 'CANCELLED'
            ),
            project_summary AS (
                SELECT 
                    project_name,
                    SUM(COALESCE(line_amount, 0)) as total_po_received,
                    SUM(CASE WHEN gap_type = 'AC_PENDING' THEN COALESCE(line_amount, 0) ELSE 0 END) as gap_po_nok_ac_nok,
                    SUM(CASE WHEN gap_type = 'PAC_PENDING' THEN COALESCE(line_amount, 0) ELSE 0 END) as gap_ac_ok_pac_nok,
                    SUM(CASE WHEN gap_type IN ('AC_PENDING', 'PAC_PENDING') THEN COALESCE(line_amount, 0) ELSE 0 END) as total_gap_ac_pac
                FROM filtered_data
                WHERE project_name IS NOT NULL AND TRIM(project_name) != ''
                GROUP BY project_name
            ),
            project_rows AS (
                SELECT 
                    project_name as "GAP by Project",
                    ROUND(total_po_received, 2) as "Total PO Received",
                    ROUND(gap_po_nok_ac_nok, 2) as "GAP PO Ok; AC Nok", 
                    ROUND(gap_ac_ok_pac_nok, 2) as "GAP AC OK; PAC Nok",
                    ROUND(total_gap_ac_pac, 2) as "Total GAP AC & PAC",
                    CASE 
                        WHEN total_po_received > 0 THEN 
                            ROUND((total_gap_ac_pac / total_po_received * 100), 0) || '%'
                        ELSE '0%'
                    END as "Pourcentage GAP Par Projet",
                    CASE 
                        WHEN total_po_received > 0 THEN 
                            (100 - ROUND((total_gap_ac_pac / total_po_received * 100), 0)) || '%'
                        ELSE '100%'
                    END as "Completion Percentage",
                    0 as sort_order
                FROM project_summary
            ),
            total_row AS (
                SELECT
                    'TOTAL' as "GAP by Project",
                    SUM("Total PO Received") as "Total PO Received",
                    SUM("GAP PO Ok; AC Nok") as "GAP PO Ok; AC Nok",
                    SUM("GAP AC OK; PAC Nok") as "GAP AC OK; PAC Nok",
                    SUM("Total GAP AC & PAC") as "Total GAP AC & PAC",
                    CASE 
                        WHEN SUM("Total PO Received") > 0 THEN 
                            ROUND((SUM("Total GAP AC & PAC") / SUM("Total PO Received") * 100), 0) || '%'
                        ELSE '0%'
                    END as "Pourcentage GAP Par Projet",
                    CASE 
                        WHEN SUM("Total PO Received") > 0 THEN 
                            (100 - ROUND((SUM("Total GAP AC & PAC") / SUM("Total PO Received") * 100), 0)) || '%'
                        ELSE '100%'
                    END as "Completion Percentage",
                    1 as sort_order
                FROM project_rows
            ),
            combined AS (
                SELECT * FROM project_rows
                UNION ALL
                SELECT * FROM total_row
            )
            SELECT 
                "GAP by Project",
                "Total PO Received",
                "GAP PO Ok; AC Nok",
                "GAP AC OK; PAC Nok",
                "Total GAP AC & PAC",
                "Pourcentage GAP Par Projet",
                "Completion Percentage"
            FROM combined
            ORDER BY sort_order, "Total PO Received" DESC
            """
            
            result = self.db.execute(text(financial_summary_query), params)
            data = result.fetchall()
            column_names = list(result.keys())
            
            # Convert to DataFrame
            df = pd.DataFrame(data, columns=column_names)
            
            # Replace dots with commas for numeric columns
            numeric_columns = ["Total PO Received", "GAP PO Ok; AC Nok", "GAP AC OK; PAC Nok", "Total GAP AC & PAC"]
            for col in numeric_columns:
                df[col] = df[col].apply(lambda x: str(x).replace('.', ','))
            
            # Create Excel output
            output = BytesIO()
            
            if df.empty:
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    empty_df = pd.DataFrame({'Message': ['No data found']})
                    empty_df.to_excel(writer, sheet_name='Gap Financial Summary', index=False)
            else:
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Gap Financial Summary', index=False)
                    
                    # Apply basic styling
                    worksheet = writer.sheets['Gap Financial Summary']
                    
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    
                    # Define styles
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    total_font = Font(bold=True)
                    total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                    
                    # Apply header formatting
                    for cell in worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    # Auto-adjust column width
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Find and format the total row
                    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                        if row[0].value == 'TOTAL':
                            for cell in row:
                                cell.font = total_font
                                cell.fill = total_fill
                                cell.alignment = Alignment(horizontal='center')
                    
            # Seek to beginning of file
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Error exporting gap financial summary to Excel: {str(e)}")
            raise