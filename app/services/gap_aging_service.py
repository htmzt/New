from typing import Dict, Any, List, Optional
from sqlalchemy.orm import Session
from sqlalchemy import text
from app.services.base_service import BaseService
from app.query import MERGED_DATA_QUERY
import pandas as pd
from io import BytesIO
import logging

logger = logging.getLogger(__name__)


class GapAgingService(BaseService):
    """Service for aging analysis of pending purchase orders"""
    
    def __init__(self, db: Session):
        super().__init__(db)
    
    def get_aging_analysis(
        self,
        user_id: str,
        project_name: Optional[str] = None,
        account_name: Optional[str] = None,
        category: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Get aging analysis of pending POs bucketed by days since publish_date
        
        Args:
            user_id: User ID to filter by
            project_name: Optional project name filter
            account_name: Optional account name filter
            category: Optional category filter
        
        Returns:
            Dictionary with aging buckets and summary statistics
        """
        try:
            # Build base filter
            base_filter = "po.user_id = :user_id"
            params = {"user_id": user_id}
            
            # Add optional filters
            if project_name:
                base_filter += " AND po.project_name ILIKE :project_name"
                params["project_name"] = f"%{project_name}%"
            
            if account_name:
                base_filter += " AND acc.account_name = :account_name"
                params["account_name"] = account_name
            
            if category:
                # Category is calculated in MERGED_DATA_QUERY, so we filter in subquery
                pass  # Will handle in WHERE clause below
            
            # Build aging analysis query
            aging_query = f"""
            WITH pending_pos AS (
                SELECT 
                    *,
                    CURRENT_DATE - publish_date AS days_old
                FROM (
                    {MERGED_DATA_QUERY.format(base_filter=base_filter)}
                ) AS merged
                WHERE 
                    -- Only pending items
                    (ac_date IS NULL OR pac_date IS NULL)
                    AND status NOT IN ('CLOSED', 'CANCELLED')
                    AND publish_date IS NOT NULL
                    -- Category filter if provided
                    {("AND category = :category" if category else "")}
            ),
            bucketed_data AS (
                SELECT 
                    CASE
                        WHEN days_old BETWEEN 0 AND 15 THEN '0-15 days'
                        WHEN days_old BETWEEN 16 AND 30 THEN '16-30 days'
                        WHEN days_old BETWEEN 31 AND 60 THEN '31-60 days'
                        WHEN days_old BETWEEN 61 AND 90 THEN '61-90 days'
                        WHEN days_old > 90 THEN '90+ days'
                        ELSE 'Unknown'
                    END AS aging_bucket,
                    
                    line_amount,
                    
                    CASE 
                        WHEN ac_date IS NULL THEN line_amount 
                        ELSE 0 
                    END AS ac_pending_amount,
                    
                    CASE 
                        WHEN ac_date IS NOT NULL AND pac_date IS NULL THEN line_amount 
                        ELSE 0 
                    END AS pac_pending_amount,
                    
                    days_old
                    
                FROM pending_pos
                WHERE days_old >= 0  -- Safety check for data quality
            )
            SELECT 
                aging_bucket,
                COUNT(*) AS po_count,
                ROUND(COALESCE(SUM(line_amount), 0), 2) AS total_amount,
                ROUND(COALESCE(SUM(ac_pending_amount), 0), 2) AS ac_pending_amount,
                ROUND(COALESCE(SUM(pac_pending_amount), 0), 2) AS pac_pending_amount,
                ROUND(AVG(days_old), 0) AS avg_days_old
            FROM bucketed_data
            GROUP BY aging_bucket
            ORDER BY 
                CASE aging_bucket
                    WHEN '0-15 days' THEN 1
                    WHEN '16-30 days' THEN 2
                    WHEN '31-60 days' THEN 3
                    WHEN '61-90 days' THEN 4
                    WHEN '90+ days' THEN 5
                    ELSE 6
                END
            """
            
            # Add category to params if provided
            if category:
                params["category"] = category
            
            # Execute query
            result = self.db.execute(text(aging_query), params)
            rows = result.fetchall()
            
            # Calculate totals and build response
            total_pos = sum(row.po_count for row in rows)
            total_amount = sum(float(row.total_amount) for row in rows)
            total_ac_pending = sum(float(row.ac_pending_amount) for row in rows)
            total_pac_pending = sum(float(row.pac_pending_amount) for row in rows)
            
            # Calculate weighted average age
            weighted_age_sum = sum(row.avg_days_old * row.po_count for row in rows)
            avg_age = round(weighted_age_sum / total_pos, 0) if total_pos > 0 else 0
            
            # Build aging buckets list
            aging_analysis = []
            for row in rows:
                percentage = round((row.po_count / total_pos) * 100, 1) if total_pos > 0 else 0
                
                bucket_data = {
                    "bucket": row.aging_bucket,
                    "po_count": row.po_count,
                    "total_amount": f"{float(row.total_amount):,.2f}".replace(',', ' ').replace('.', ','),
                    "ac_pending_amount": f"{float(row.ac_pending_amount):,.2f}".replace(',', ' ').replace('.', ','),
                    "pac_pending_amount": f"{float(row.pac_pending_amount):,.2f}".replace(',', ' ').replace('.', ','),
                    "percentage": f"{percentage}%",
                    "avg_days_old": int(row.avg_days_old),
                    "status": "Pending"
                }
                
                # Add alert level for UI
                if row.aging_bucket == '0-15 days':
                    bucket_data["alert_level"] = "success"
                elif row.aging_bucket == '16-30 days':
                    bucket_data["alert_level"] = "info"
                elif row.aging_bucket == '31-60 days':
                    bucket_data["alert_level"] = "warning"
                elif row.aging_bucket == '61-90 days':
                    bucket_data["alert_level"] = "danger"
                elif row.aging_bucket == '90+ days':
                    bucket_data["alert_level"] = "critical"
                
                aging_analysis.append(bucket_data)
            
            return {
                "success": True,
                "filters_applied": {
                    "project_name": project_name,
                    "account_name": account_name,
                    "category": category
                },
                "summary": {
                    "total_pending_pos": total_pos,
                    "total_pending_amount": f"{total_amount:,.2f}".replace(',', ' ').replace('.', ','),
                    "total_ac_pending": f"{total_ac_pending:,.2f}".replace(',', ' ').replace('.', ','),
                    "total_pac_pending": f"{total_pac_pending:,.2f}".replace(',', ' ').replace('.', ','),
                    "average_age_days": int(avg_age)
                },
                "aging_analysis": aging_analysis
            }
            
        except Exception as e:
            logger.error(f"Error in aging analysis: {str(e)}")
            raise
    
    def export_aging_analysis_to_excel(
        self,
        user_id: str,
        project_name: Optional[str] = None,
        account_name: Optional[str] = None,
        category: Optional[str] = None
    ) -> bytes:
        """
        Export aging analysis to Excel file
        
        Args:
            user_id: User ID to filter by
            project_name: Optional project name filter
            account_name: Optional account name filter
            category: Optional category filter
        
        Returns:
            Excel file as bytes
        """
        try:
            # Get aging analysis data
            analysis_data = self.get_aging_analysis(
                user_id=user_id,
                project_name=project_name,
                account_name=account_name,
                category=category
            )
            
            if not analysis_data["aging_analysis"]:
                # Create empty Excel with message
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    empty_df = pd.DataFrame({'Message': ['No pending data found']})
                    empty_df.to_excel(writer, sheet_name='Aging Analysis', index=False)
                output.seek(0)
                return output.getvalue()
            
            # Prepare data for Excel
            excel_data = []
            for bucket in analysis_data["aging_analysis"]:
                excel_data.append({
                    "Aging Bucket": bucket["bucket"],
                    "PO Count": bucket["po_count"],
                    "Total Amount": bucket["total_amount"],
                    "AC Pending": bucket["ac_pending_amount"],
                    "PAC Pending": bucket["pac_pending_amount"],
                    "% of Total": bucket["percentage"],
                    "Avg Days Old": bucket["avg_days_old"],
                    "Status": bucket["status"]
                })
            
            # Add summary row
            excel_data.append({
                "Aging Bucket": "TOTAL",
                "PO Count": analysis_data["summary"]["total_pending_pos"],
                "Total Amount": analysis_data["summary"]["total_pending_amount"],
                "AC Pending": analysis_data["summary"]["total_ac_pending"],
                "PAC Pending": analysis_data["summary"]["total_pac_pending"],
                "% of Total": "100%",
                "Avg Days Old": analysis_data["summary"]["average_age_days"],
                "Status": "Pending"
            })
            
            # Create DataFrame
            df = pd.DataFrame(excel_data)
            
            # Create Excel file
            output = BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Aging Analysis', index=False)
                
                # Apply styling
                worksheet = writer.sheets['Aging Analysis']
                
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                # Define styles
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                total_font = Font(bold=True)
                total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                
                # Color coding for aging buckets
                success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
                info_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")      # Blue
                warning_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")   # Yellow
                danger_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")    # Orange
                critical_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Red
                
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Style header row
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = border
                
                # Style data rows
                total_row_num = len(df) + 1
                for row_num in range(2, len(df) + 2):
                    bucket_value = worksheet.cell(row=row_num, column=1).value
                    
                    for col_num in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.border = border
                        
                        # Apply bold and gray background to total row
                        if row_num == total_row_num:
                            cell.font = total_font
                            cell.fill = total_fill
                        # Color code by aging bucket
                        elif col_num == 1 and bucket_value != "TOTAL":
                            if "0-15" in str(bucket_value):
                                cell.fill = success_fill
                            elif "16-30" in str(bucket_value):
                                cell.fill = info_fill
                            elif "31-60" in str(bucket_value):
                                cell.fill = warning_fill
                            elif "61-90" in str(bucket_value):
                                cell.fill = danger_fill
                            elif "90+" in str(bucket_value):
                                cell.fill = critical_fill
                        
                        # Alignment
                        if col_num == 1 or col_num == len(df.columns):  # First and last columns
                            cell.alignment = Alignment(horizontal="left")
                        else:
                            cell.alignment = Alignment(horizontal="right")
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add a summary sheet
                summary_data = [
                    {"Metric": "Total Pending POs", "Value": analysis_data["summary"]["total_pending_pos"]},
                    {"Metric": "Total Pending Amount", "Value": analysis_data["summary"]["total_pending_amount"]},
                    {"Metric": "AC Pending Amount", "Value": analysis_data["summary"]["total_ac_pending"]},
                    {"Metric": "PAC Pending Amount", "Value": analysis_data["summary"]["total_pac_pending"]},
                    {"Metric": "Average Age (Days)", "Value": analysis_data["summary"]["average_age_days"]},
                ]
                
                if project_name:
                    summary_data.insert(0, {"Metric": "Project Filter", "Value": project_name})
                if account_name:
                    summary_data.insert(0, {"Metric": "Account Filter", "Value": account_name})
                if category:
                    summary_data.insert(0, {"Metric": "Category Filter", "Value": category})
                
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Style summary sheet
                summary_sheet = writer.sheets['Summary']
                for col_num in range(1, 3):
                    cell = summary_sheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = border
                
                for row_num in range(2, len(summary_data) + 2):
                    for col_num in range(1, 3):
                        cell = summary_sheet.cell(row=row_num, column=col_num)
                        cell.border = border
                
                summary_sheet.column_dimensions['A'].width = 25
                summary_sheet.column_dimensions['B'].width = 30
            
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Error exporting aging analysis to Excel: {str(e)}")
            raise